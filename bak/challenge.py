# -*- coding: utf-8 -*-

from openpyxl import load_workbook 
from openpyxl import Workbook 
import datetime

def combine():
    wb=load_workbook('/home/shiyanlou/Code/courses.xlsx')
    ws1=wb.get_sheet_by_name('students')
    ws2=wb.get_sheet_by_name('time')
    ws3=wb.create_sheet('combine')
    a=[x.value for x in ws1['A']]
    b=[x.value for x in ws1['B']]
    c=[x.value for x in ws1['C']]
    d=[x.value for x in ws2['B']]
    e=[x.value for x in ws2['C']]
    f=[]
    for x in b:
        m=d.index(x)
        f.append(e[m])


    for i in range(1,len(a)+1):
        ws3['A'+str(i)]=a[i-1]
    for i in range(1,len(b)+1):
        ws3['B'+str(i)]=b[i-1]
    for i in range(1,len(c)+1):
        ws3['C'+str(i)]=c[i-1]
    for i in range(1,len(d)+1):
        ws3['D'+str(i)]=f[i-1]

    wb.save('/home/shiyanlou/Code/courses.xlsx')


def split():
    pass


if __name__ == '__main__':
    combine()
    split()
